import csv
import os
import openpyxl

day = "20220724"
txt = "./QBAnalysis/" + day + ".txt"
csvg = "./QBAnalysis/" + day + ".csv"
xlsx = "./QBAnalysis/data.xlsx"
# txt = day + ".txt"
# csvg = day + ".csv"
# xlsx = "./data.xlsx"

# if os.path.exists(xlsx):
#   print('yes')
# else:
#   print('no')
with open(txt, mode='r') as f:
  s = f.read()
  s = s.replace('</g>', '</g>\n')
  s = s.replace('<g data-v-eb0ea46c=""><g data-v-eb0ea46c=""><circle data-v-eb0ea46c=""', '')
  s = s.replace('<g data-v-eb0ea46c=""><circle data-v-eb0ea46c="" ', '')
  s = s.replace(' r="3" class="graph__data-circle--cbt"></circle></g>', '')
  s = s.replace(' r="5" class="graph__data-circle--cbt"></circle></g>', '')
  s = s.replace('</g>', '')
  s = s.replace('\n<g data-v-eb0ea46c=""><!---->', '')
  s = s.replace(' r="3" class="graph__my-data-circle--highlighted"></circle>', ' 自分')
  s = s.replace(' r="5" class="graph__my-data-circle--highlighted"></circle>', ' 自分')
  s = s.replace('cx="', '')
  s = s.replace('" cy="', ',')
  s = s.replace('"', ',')
  s = s.replace('\n\n', ',')
  with open(csvg, mode='w') as f1: 
    f1.write(s)

#CSVファイルを開く
with open(csvg,newline="") as csvf:
  #CSVファイルを読み込む
  data = csv.reader(csvf)
  #Excelファイルを開く
  if not os.path.exists(xlsx):
    wa = openpyxl.Workbook()
    wa.save(filename=xlsx)
    wa.close()
  wb = openpyxl.load_workbook(filename=xlsx)
  #sheetを読み込む
  wb.create_sheet(title=day)
  r = 1
  for line in data:
    c = 1
    for v in line:
      wb[day].cell(row=r,column=c).value=v
      c += 1
    r += 1
#ファイルを保存する
wb.save(filename=xlsx)
#ファイルを閉じる
wb.close()
os.remove(csvg)
