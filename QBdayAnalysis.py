import csv
import openpyxl
import os

day = "629712"
txt = "./QBAnalysis/" + day + ".txt"
txtmine = "./QBAnalysis/" + day + "mine.txt"
csvg = "./QBAnalysis/" + day + ".csv"
csvgmine = "./QBAnalysis/" + day + "mine.csv"
xlsx = "./QBAnalysis/daydata.xlsx"

#置換する文字列を指定
Henkan_mae = ['122','154','186','218','250','282','314','346','378','410','442','474','506','538']
Henkan_go = ['7/13','7/14','7/15','7/16','7/17','7/18','7/19','7/20','7/21','7/22','7/23','7/24','7/25','7/26']
# txt = day + ".txt"
# csvg = day + ".csv"
# xlsx = "./data.xlsx"

# if os.path.exists(xlsx):
#   print('yes')
# else:
#   print('no')
with open(txt, mode='r') as f:
  s = f.read()
  s = s.replace('</g>', '</g>\n')
  s = s.replace('<g data-v-20bde5c5=""><g data-v-20bde5c5=""><circle data-v-20bde5c5="" ', '')
  s = s.replace('<g data-v-20bde5c5=""><circle data-v-20bde5c5="" ', '')
  s = s.replace(' r="5"', '')
  s = s.replace(' class="graph__circle--cbt"></circle></g>', '')
  s = s.replace(' class="graph__circle--my-data"></circle></g>', '1')
  s = s.replace('\n<g data-v-20bde5c5=""><!----></g>', '')
  s = s.replace('</g>', '')
  s = s.replace('cx="', '')
  s = s.replace('" cy="', ',')
  s = s.replace('" data-count="', ',')
  s = s.replace('"', ',')
  s = s.replace('\n\n', ',')
  # print(s)
  with open(csvg, mode='w') as f1: 
    f1.write(s)

with open(txtmine, mode='r') as f:
  s = f.read()
  s = s.replace('</g>', '</g>\n')
  s = s.replace('<g data-v-20bde5c5=""><g data-v-20bde5c5=""><circle data-v-20bde5c5="" ', '')
  s = s.replace('<g data-v-20bde5c5=""><circle data-v-20bde5c5="" ', '')
  s = s.replace(' r="5"', '')
  s = s.replace(' class="graph__circle--cbt"></circle></g>', '')
  s = s.replace(' class="graph__circle--my-data"></circle></g>', '自分')
  s = s.replace('\n<g data-v-20bde5c5=""><!----></g>', '')
  s = s.replace('</g>', '')
  s = s.replace('cx="', '')
  s = s.replace('" cy="', ',')
  s = s.replace('" data-count="', ',')
  s = s.replace('"', ',')
  s = s.replace('\n\n', ',')
  # print(s)
  with open(csvgmine, mode='w') as f1: 
    f1.write(s)

#CSVファイルを開く
with open(csvg,newline="") as csvf:
  with open(csvgmine,newline="") as csvfmine:
    #CSVファイルを読み込む
    data = csv.reader(csvf)
    datamine = csv.reader(csvfmine)
    #Excelファイルを開く
    if not os.path.exists(xlsx):
      wa = openpyxl.Workbook()
      wa.save(filename=xlsx)
      wa.close()
    wb = openpyxl.load_workbook(filename=xlsx)
    #sheetを読み込む
    wb.create_sheet(title=day)
    r = 1
    for line in data:
      c = 1
      for v in line:
        wb[day].cell(row=r,column=c).value=v
        c += 1
      r += 1
    for line in datamine:
      c = 1
      for v in line:
        wb[day].cell(row=r,column=c).value=v
        c += 1
      r += 1
    ws = wb[day]
    ws.delete_cols(2)
    i = 0
    #リストをループ
    for list in Henkan_mae:
      i = i + 1
      #セルをループ
      for row in ws.iter_rows():
        for cell in row:
          #1列目だったら
          if cell.col_idx == 1:
            #セルにリストが含まれていたら
            if list in cell.value:
              #置換
              new_text = cell.value.replace(list, Henkan_go[i-1])
              cell.value = new_text
#ファイルを保存する
wb.save(filename=xlsx)
#ファイルを閉じる
wb.close()
os.remove(csvg)
os.remove(csvgmine)